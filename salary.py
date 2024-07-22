import openpyxl
import datetime
from os.path import getctime

months = {
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь'
}


def get_salary_report(message, employee, full_report):
    fl_name_split = employee[message].split('_')
    full_name = fl_name_split[1] + ' ' + fl_name_split[0]
    print(f'getting salary info for {full_name}...\nfull_report = {full_report}')
    now = datetime.datetime.now()
    now_month = now.month
    wb = openpyxl.load_workbook('files/зп.xlsx', data_only=True)
    sheet = wb[months[now_month]]   #удалить +1
    print(sheet)
    report = ''
    try:
        for i in range(3, 30):
            if sheet.cell(row=i, column=1).value == full_name and sheet.cell(row=i, column=1).value != sheet.cell(row=69, column=69).value and sheet.cell(row=i, column=1).value != 'Расходы на такси' and\
                    sheet.cell(row=i, column=1).value != 'Хоз. Работы':
                print(f'row number finded - {i}')
                staff_row = i
        hours_monfri = 0
        hours_satsun = 0
        hours_voc_150 = 0
        hours_voc_170 = 0
        hours_voc_180 = 0
        hours_monfrihot = 0
        hours_satsunhot = 0
        total = 0
        salary = 0
        for i in range(2, 60):
            match sheet.cell(row=1, column=i).value:
                case 'будние':
                    hours_monfri = sheet.cell(row=staff_row, column=i).value
                case 'выходные':
                    hours_satsun = sheet.cell(row=staff_row, column=i).value
                case 'праздники(150)':
                    hours_voc_150 = sheet.cell(row=staff_row, column=i).value
                    print(hours_voc_150)
                case 'праздники(170)':
                    hours_voc_170 = sheet.cell(row=staff_row, column=i).value
                    print(hours_voc_170)
                case 'праздники(180)':
                    hours_voc_180 = sheet.cell(row=staff_row, column=i).value
                    print(hours_voc_180)
                case 'будние +30':
                    hours_monfrihot = sheet.cell(row=staff_row, column=i).value
                case 'выходные +30':
                    hours_satsunhot = sheet.cell(row=staff_row, column=i).value
                case 'итого часов':
                    total = sheet.cell(row=staff_row, column=i).value
                case 'зарплата':
                    salary = sheet.cell(row=staff_row, column=i).value
                case _:
                    if str(sheet.cell(row=2, column=i).value) != 'None' and str(sheet.cell(row=staff_row, column=i).value) != 'None' and full_report is True:
                        report += str(sheet.cell(row=1, column=i).value) + ', ' + str(sheet.cell(row=2, column=i).value) + ' - ' + str(sheet.cell(row=staff_row, column=i).value) + '\n'
        if full_report is True:
            mtime_readable = datetime.datetime.fromtimestamp(getctime('files/зп.xlsx')).strftime("%d.%m.%Y г. %H:%M")
            return f'🎢Полный отчет для: {full_name}🎢\nПоследнее обновление таблицы: {mtime_readable}\n\n{report}\n\n🕑Будние: {hours_monfri} ч.\n🕑Выходные: {hours_satsun} ч.\n\n🕑Праздники(150): {hours_voc_150} ч.\n🕑Праздники(170): {hours_voc_170} ч.\n🕑Праздники(180): {hours_voc_180} ч.\n\n🕑Будние(>+30°C): {hours_monfrihot} ч.\n🕑Выходные(>30°C): {hours_satsunhot} ч.\n\n🥳ИТОГО: {total} ч.\n\n💰Зарплата: {salary}₽'
        else:
            mtime_readable = datetime.datetime.fromtimestamp(getctime('files/зп.xlsx')).strftime("%d.%m.%Y г. %H:%M")
            return f'🎢Зарплата для: {full_name}🎢\nПоследнее обновление таблицы: {mtime_readable}\n\n🕑Будние: {hours_monfri} ч.\n🕑Выходные: {hours_satsun} ч.\n\n🕑Праздники(150): {hours_voc_150} ч.\n🕑Праздники(170): {hours_voc_170} ч.\n🕑Праздники(180): {hours_voc_180} ч.\n\n🕑Будние(>+30°C): {hours_monfrihot} ч.\n🕑Выходные(>30°C): {hours_satsunhot} ч.\n\n🥳ИТОГО: {total} ч.\n\n💰Зарплата: {salary}₽'
    except Exception as e:
        print(e)
        return f'Произошла какая-то ошибка... Я не нашел никаких данных для {full_name}, пожалуйста, обратись к администратору.'
